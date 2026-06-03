"""
Purpose: Parse and validate imported general-ledger and trial-balance upload files.
Scope: CSV/XLSX/searchable-PDF decoding, canonical header normalization, amount/date validation,
and conversion into typed import seeds for service-layer persistence.
Dependencies: Python CSV/io helpers, Decimal/date parsing, and openpyxl workbook reads.
"""

from __future__ import annotations

import csv
import hashlib
import re
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from enum import StrEnum
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]
from pypdf import PdfReader
from pypdf.errors import PdfReadError
from services.common.types import JsonObject
from services.imports.intelligence import (
    IMPORT_INTELLIGENCE_METADATA_KEY,
    ImportColumnMapping,
    ImportDiagnosticIssue,
    build_header_column_mappings,
    build_import_intelligence_report,
)


class LedgerImportErrorCode(StrEnum):
    """Enumerate stable validation codes surfaced by ledger baseline uploads."""

    INVALID_FILE = "invalid_file"
    UNSUPPORTED_FILE_TYPE = "unsupported_file_type"


class LedgerImportError(ValueError):
    """Represent a fail-fast ledger baseline validation failure."""

    def __init__(self, *, code: LedgerImportErrorCode, message: str) -> None:
        super().__init__(message)
        self.code = code
        self.message = message


@dataclass(frozen=True, slots=True)
class ImportedGeneralLedgerLineSeed:
    """Describe one validated imported general-ledger line."""

    line_no: int
    posting_date: date
    account_code: str
    account_name: str | None
    reference: str | None
    description: str | None
    debit_amount: Decimal
    credit_amount: Decimal
    dimensions: JsonObject
    external_ref: str | None
    transaction_group_key: str


@dataclass(frozen=True, slots=True)
class ImportedTrialBalanceLineSeed:
    """Describe one validated imported trial-balance row."""

    line_no: int
    account_code: str
    account_name: str | None
    account_type: str | None
    debit_balance: Decimal
    credit_balance: Decimal
    is_active: bool


@dataclass(frozen=True, slots=True)
class ImportedGeneralLedgerFile:
    """Describe one validated general-ledger import payload."""

    lines: tuple[ImportedGeneralLedgerLineSeed, ...]
    import_metadata: JsonObject


@dataclass(frozen=True, slots=True)
class ImportedTrialBalanceFile:
    """Describe one validated trial-balance import payload."""

    lines: tuple[ImportedTrialBalanceLineSeed, ...]
    import_metadata: JsonObject


@dataclass(frozen=True, slots=True)
class _CanonicalRow:
    """Describe one canonicalized source row with its original 1-based row number."""

    values: dict[str, str]
    row_number: int


@dataclass(frozen=True, slots=True)
class _CanonicalRowsResult:
    """Describe canonicalized rows plus header-detection evidence."""

    rows: tuple[_CanonicalRow, ...]
    detected_columns: frozenset[str]
    header_row_index: int
    column_mappings: tuple[ImportColumnMapping, ...]


_GL_REQUIRED_COLUMNS = frozenset({"posting_date"})
_TB_REQUIRED_COLUMNS = frozenset[str]()
_SUPPORTED_FORMATS = ("csv", "xlsx", "pdf")

_GL_HEADER_ALIASES = {
    "account": "account_name",
    "account_code": "account_code",
    "account_description": "account_name",
    "account_name": "account_name",
    "account_number": "account_code",
    "account_title": "account_name",
    "amount": "amount",
    "cost_centre": "cost_centre",
    "cost_center": "cost_centre",
    "credit": "credit_amount",
    "credit_amount": "credit_amount",
    "date": "posting_date",
    "department": "department",
    "description": "description",
    "debit": "debit_amount",
    "debit_amount": "debit_amount",
    "entry_id": "transaction_group_key",
    "entry_key": "transaction_group_key",
    "entry_no": "transaction_group_key",
    "entry_number": "transaction_group_key",
    "external_ref": "external_ref",
    "external_reference": "external_ref",
    "gl_code": "account_code",
    "group_id": "transaction_group_key",
    "group_key": "transaction_group_key",
    "journal_date": "posting_date",
    "journal_id": "transaction_group_key",
    "journal_key": "transaction_group_key",
    "journal_no": "transaction_group_key",
    "journal_number": "transaction_group_key",
    "line_type": "line_type",
    "memo": "description",
    "name": "account_name",
    "posting_date": "posting_date",
    "project": "project",
    "ref": "reference",
    "reference": "reference",
    "signed_amount": "signed_amount",
    "transaction_date": "posting_date",
    "transaction_group": "transaction_group_key",
    "transaction_group_id": "transaction_group_key",
    "transaction_group_key": "transaction_group_key",
    "transaction_id": "transaction_group_key",
    "transaction_key": "transaction_group_key",
    "transaction_no": "transaction_group_key",
    "transaction_number": "transaction_group_key",
    "transaction_ref": "reference",
    "type": "line_type",
    "voucher_id": "transaction_group_key",
    "voucher_no": "transaction_group_key",
    "voucher_number": "transaction_group_key",
}

_TB_HEADER_ALIASES = {
    "account": "account_name",
    "account_code": "account_code",
    "account_description": "account_name",
    "account_name": "account_name",
    "account_number": "account_code",
    "account_title": "account_name",
    "account_type": "account_type",
    "active": "is_active",
    "balance": "balance",
    "balance_side": "balance_side",
    "balance_type": "balance_side",
    "code": "account_code",
    "credit": "credit_balance",
    "credit_balance": "credit_balance",
    "debit": "debit_balance",
    "debit_balance": "debit_balance",
    "gl_code": "account_code",
    "is_active": "is_active",
    "name": "account_name",
    "status": "is_active",
    "type": "account_type",
}

_TRUE_LITERALS = frozenset({"1", "active", "t", "true", "y", "yes"})
_FALSE_LITERALS = frozenset({"0", "f", "false", "inactive", "n", "no"})
_DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d")


def import_general_ledger_file(
    *,
    filename: str,
    payload: bytes,
    account_code_lookup: Mapping[str, str] | None = None,
) -> ImportedGeneralLedgerFile:
    """Parse a CSV/XLSX/searchable-PDF general-ledger payload into line seeds."""

    read_result, source_format = _read_rows(
        filename=filename,
        payload=payload,
        header_aliases=_GL_HEADER_ALIASES,
        required_columns=_GL_REQUIRED_COLUMNS,
        required_any_column_groups=(frozenset({"account_code", "account_name"}),),
        noun="general ledger",
    )
    normalized_account_lookup = _normalize_account_code_lookup(account_code_lookup)
    lines = tuple(
        _parse_gl_row(
            row=row.values,
            row_number=row.row_number,
            account_code_lookup=normalized_account_lookup,
        )
        for row in read_result.rows
    )
    account_identity_strategy = (
        "explicit_account_code"
        if "account_code" in read_result.detected_columns
        else "resolved_by_name"
    )
    transaction_grouping_strategy = (
        "explicit_column"
        if "transaction_group_key" in read_result.detected_columns
        else "derived_from_ledger_fields"
    )
    warnings = _build_import_warnings(
        document_kind="general_ledger",
        source_format=source_format,
        account_identity_strategy=account_identity_strategy,
        transaction_grouping_strategy=transaction_grouping_strategy,
    )
    metadata: JsonObject = {
        "detected_columns": ", ".join(sorted(read_result.detected_columns)),
        "format": source_format,
        "row_count": len(lines),
        "account_identity_strategy": account_identity_strategy,
        "transaction_grouping_strategy": transaction_grouping_strategy,
        "uploaded_filename": filename,
        IMPORT_INTELLIGENCE_METADATA_KEY: build_import_intelligence_report(
            document_kind="general_ledger",
            source_format=source_format,
            uploaded_filename=filename,
            row_count=len(read_result.rows),
            accepted_row_count=len(lines),
            detected_columns=tuple(read_result.detected_columns),
            header_row_index=read_result.header_row_index,
            column_mappings=read_result.column_mappings,
            confidence=_resolve_import_confidence(
                source_format=source_format,
                account_identity_strategy=account_identity_strategy,
                transaction_grouping_strategy=transaction_grouping_strategy,
            ),
            parsing_strategy="header_row_scan_with_alias_mapping",
            parser_capabilities=_SUPPORTED_FORMATS,
            warnings=warnings,
            recovery_actions=tuple(
                warning.recovery_action
                for warning in warnings
                if warning.recovery_action is not None
            ),
            extra={
                "account_identity_strategy": account_identity_strategy,
                "account_name_resolution_count": _count_account_name_resolution_rows(
                    rows=read_result.rows,
                ),
                "amount_strategy": _detect_gl_amount_strategy(read_result.detected_columns),
                "transaction_grouping_strategy": transaction_grouping_strategy,
            },
        ),
    }
    return ImportedGeneralLedgerFile(lines=lines, import_metadata=metadata)


def import_trial_balance_file(
    *,
    filename: str,
    payload: bytes,
    account_code_lookup: Mapping[str, str] | None = None,
) -> ImportedTrialBalanceFile:
    """Parse a CSV/XLSX/searchable-PDF trial-balance payload into account seeds."""

    read_result, source_format = _read_rows(
        filename=filename,
        payload=payload,
        header_aliases=_TB_HEADER_ALIASES,
        required_columns=_TB_REQUIRED_COLUMNS,
        required_any_column_groups=(frozenset({"account_code", "account_name"}),),
        noun="trial balance",
    )
    normalized_account_lookup = _normalize_account_code_lookup(account_code_lookup)
    data_rows = tuple(
        row for row in read_result.rows if not _is_trial_balance_summary_row(row=row.values)
    )
    lines = tuple(
        _parse_tb_row(
            row=row.values,
            row_number=row.row_number,
            account_code_lookup=normalized_account_lookup,
        )
        for row in data_rows
    )
    if not lines:
        raise LedgerImportError(
            code=LedgerImportErrorCode.INVALID_FILE,
            message="The trial balance file does not contain any account rows.",
        )
    account_identity_strategy = (
        "explicit_account_code"
        if "account_code" in read_result.detected_columns
        else "resolved_by_name"
    )
    skipped_summary_row_count = len(read_result.rows) - len(data_rows)
    warnings = _build_import_warnings(
        document_kind="trial_balance",
        source_format=source_format,
        account_identity_strategy=account_identity_strategy,
        skipped_summary_row_count=skipped_summary_row_count,
    )
    metadata: JsonObject = {
        "detected_columns": ", ".join(sorted(read_result.detected_columns)),
        "format": source_format,
        "row_count": len(lines),
        "account_identity_strategy": account_identity_strategy,
        "uploaded_filename": filename,
        IMPORT_INTELLIGENCE_METADATA_KEY: build_import_intelligence_report(
            document_kind="trial_balance",
            source_format=source_format,
            uploaded_filename=filename,
            row_count=len(read_result.rows),
            accepted_row_count=len(lines),
            detected_columns=tuple(read_result.detected_columns),
            header_row_index=read_result.header_row_index,
            column_mappings=read_result.column_mappings,
            confidence=_resolve_import_confidence(
                source_format=source_format,
                account_identity_strategy=account_identity_strategy,
                transaction_grouping_strategy=None,
            ),
            parsing_strategy="header_row_scan_with_alias_mapping",
            parser_capabilities=_SUPPORTED_FORMATS,
            warnings=warnings,
            recovery_actions=tuple(
                warning.recovery_action
                for warning in warnings
                if warning.recovery_action is not None
            ),
            extra={
                "account_identity_strategy": account_identity_strategy,
                "account_name_resolution_count": _count_account_name_resolution_rows(
                    rows=data_rows,
                ),
                "balance_strategy": _detect_tb_balance_strategy(read_result.detected_columns),
                "skipped_summary_row_count": skipped_summary_row_count,
            },
        ),
    }
    return ImportedTrialBalanceFile(lines=lines, import_metadata=metadata)


def _read_rows(
    *,
    filename: str,
    payload: bytes,
    header_aliases: dict[str, str],
    required_columns: frozenset[str],
    required_any_column_groups: tuple[frozenset[str], ...] = (),
    noun: str,
) -> tuple[_CanonicalRowsResult, str]:
    """Read canonicalized rows from a CSV, XLSX, or searchable PDF payload."""

    if not payload:
        raise LedgerImportError(
            code=LedgerImportErrorCode.INVALID_FILE,
            message=f"Uploaded {noun} files cannot be empty.",
        )

    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return (
            _read_csv_rows(
                payload=payload,
                header_aliases=header_aliases,
                required_columns=required_columns,
                required_any_column_groups=required_any_column_groups,
                noun=noun,
            ),
            "csv",
        )
    if suffix in {".xlsx", ".xlsm"}:
        return (
            _read_workbook_rows(
                payload=payload,
                header_aliases=header_aliases,
                required_columns=required_columns,
                required_any_column_groups=required_any_column_groups,
                noun=noun,
            ),
            "xlsx",
        )
    if suffix == ".pdf":
        return (
            _read_pdf_rows(
                payload=payload,
                header_aliases=header_aliases,
                required_columns=required_columns,
                required_any_column_groups=required_any_column_groups,
                noun=noun,
            ),
            "pdf",
        )

    raise LedgerImportError(
        code=LedgerImportErrorCode.UNSUPPORTED_FILE_TYPE,
        message=f"Upload a CSV, XLSX, or searchable PDF {noun} file.",
    )


def _read_csv_rows(
    *,
    payload: bytes,
    header_aliases: dict[str, str],
    required_columns: frozenset[str],
    required_any_column_groups: tuple[frozenset[str], ...],
    noun: str,
) -> _CanonicalRowsResult:
    """Read CSV rows and normalize headers into canonical column names."""

    try:
        decoded_payload = payload.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise LedgerImportError(
            code=LedgerImportErrorCode.INVALID_FILE,
            message="CSV files must be UTF-8 encoded.",
        ) from error

    try:
        dialect = csv.Sniffer().sniff(decoded_payload[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(StringIO(decoded_payload), dialect=dialect)
    raw_rows = tuple(tuple(cell.strip() for cell in row) for row in reader)
    if not raw_rows:
        raise LedgerImportError(
            code=LedgerImportErrorCode.INVALID_FILE,
            message=f"The {noun} CSV file must include a header row.",
        )

    return _read_matrix_rows(
        raw_rows=raw_rows,
        header_aliases=header_aliases,
        required_columns=required_columns,
        required_any_column_groups=required_any_column_groups,
        noun=noun,
    )


def _read_workbook_rows(
    *,
    payload: bytes,
    header_aliases: dict[str, str],
    required_columns: frozenset[str],
    required_any_column_groups: tuple[frozenset[str], ...],
    noun: str,
) -> _CanonicalRowsResult:
    """Read the first worksheet in an XLSX payload and normalize headers."""

    try:
        workbook = load_workbook(filename=BytesIO(payload), read_only=True, data_only=True)
    except Exception as error:
        raise LedgerImportError(
            code=LedgerImportErrorCode.INVALID_FILE,
            message="The workbook could not be opened. Upload a valid XLSX file.",
        ) from error

    worksheet = workbook.active
    raw_rows = tuple(
        tuple("" if cell is None else str(cell).strip() for cell in row)
        for row in worksheet.iter_rows(values_only=True)
    )
    if not raw_rows:
        raise LedgerImportError(
            code=LedgerImportErrorCode.INVALID_FILE,
            message=f"The {noun} workbook does not contain a header row.",
        )

    return _read_matrix_rows(
        raw_rows=raw_rows,
        header_aliases=header_aliases,
        required_columns=required_columns,
        required_any_column_groups=required_any_column_groups,
        noun=noun,
    )


def _read_pdf_rows(
    *,
    payload: bytes,
    header_aliases: dict[str, str],
    required_columns: frozenset[str],
    required_any_column_groups: tuple[frozenset[str], ...],
    noun: str,
) -> _CanonicalRowsResult:
    """Read searchable PDF text when the table structure is preserved by extraction."""

    try:
        reader = PdfReader(BytesIO(payload))
    except PdfReadError as error:
        raise LedgerImportError(
            code=LedgerImportErrorCode.INVALID_FILE,
            message="The PDF could not be opened. Upload a valid searchable PDF.",
        ) from error

    raw_rows = tuple(
        tuple(cell.strip() for cell in line.split("|"))
        for page in reader.pages
        for line in (page.extract_text() or "").splitlines()
        if "|" in line
    )
    if not raw_rows:
        raise LedgerImportError(
            code=LedgerImportErrorCode.INVALID_FILE,
            message=(
                f"The PDF text did not preserve a {noun} table. Upload the source spreadsheet "
                "as XLSX/CSV, or a searchable PDF whose table columns extract cleanly."
            ),
        )

    return _read_matrix_rows(
        raw_rows=raw_rows,
        header_aliases=header_aliases,
        required_columns=required_columns,
        required_any_column_groups=required_any_column_groups,
        noun=noun,
    )


def _build_header_map(
    *,
    headers: Sequence[str],
    header_aliases: dict[str, str],
    required_columns: frozenset[str],
    required_any_column_groups: tuple[frozenset[str], ...],
    noun: str,
) -> dict[int, str]:
    """Map source headers to canonical field names and validate required columns."""

    header_map: dict[int, str] = {}
    for index, header in enumerate(headers):
        normalized = _normalize_header_name(header)
        canonical_name = header_aliases.get(normalized)
        if canonical_name is None:
            continue
        header_map[index] = canonical_name

    missing = sorted(required_columns.difference(header_map.values()))
    if missing:
        missing_columns = ", ".join(missing)
        raise LedgerImportError(
            code=LedgerImportErrorCode.INVALID_FILE,
            message=f"The {noun} file is missing required columns: {missing_columns}.",
        )
    detected_columns = set(header_map.values())
    for required_group in required_any_column_groups:
        if detected_columns.isdisjoint(required_group):
            options = " or ".join(sorted(required_group))
            raise LedgerImportError(
                code=LedgerImportErrorCode.INVALID_FILE,
                message=f"The {noun} file must include {options}.",
            )

    return header_map


def _read_matrix_rows(
    *,
    raw_rows: Sequence[Sequence[str]],
    header_aliases: dict[str, str],
    required_columns: frozenset[str],
    required_any_column_groups: tuple[frozenset[str], ...],
    noun: str,
) -> _CanonicalRowsResult:
    """Find the most likely header row and canonicalize rows below it."""

    indexed_non_empty_rows = tuple(
        (index, row) for index, row in enumerate(raw_rows) if any(cell.strip() for cell in row)
    )
    if not indexed_non_empty_rows:
        raise LedgerImportError(
            code=LedgerImportErrorCode.INVALID_FILE,
            message=f"The {noun} file does not contain any data rows.",
        )

    best_header: tuple[int, int, dict[int, str]] | None = None
    best_score = -1
    for position, (row_index, row) in enumerate(indexed_non_empty_rows[:25]):
        header_map = {
            column_index: header_aliases[normalized]
            for column_index, header in enumerate(row)
            if (normalized := _normalize_header_name(header)) in header_aliases
        }
        score = _score_header_map(header_map)
        if score > best_score:
            best_header = (position, row_index, header_map)
            best_score = score

    if best_header is None:
        _build_header_map(
            headers=(),
            header_aliases=header_aliases,
            required_columns=required_columns,
            required_any_column_groups=required_any_column_groups,
            noun=noun,
        )
    assert best_header is not None

    header_position, header_row_index, header_map = best_header
    header_map = _build_header_map(
        headers=indexed_non_empty_rows[header_position][1],
        header_aliases=header_aliases,
        required_columns=required_columns,
        required_any_column_groups=required_any_column_groups,
        noun=noun,
    )
    headers = tuple(indexed_non_empty_rows[header_position][1])
    rows: list[_CanonicalRow] = []
    for row_index, raw_row in indexed_non_empty_rows[header_position + 1 :]:
        canonical = _canonicalize_matrix_row(raw_row=raw_row, header_map=header_map)
        if canonical:
            rows.append(_CanonicalRow(values=canonical, row_number=row_index + 1))
    if not rows:
        raise LedgerImportError(
            code=LedgerImportErrorCode.INVALID_FILE,
            message=f"The {noun} file does not contain any data rows.",
        )
    return _CanonicalRowsResult(
        rows=tuple(rows),
        detected_columns=frozenset(header_map.values()),
        header_row_index=header_row_index,
        column_mappings=build_header_column_mappings(
            headers=headers,
            header_map=header_map,
        ),
    )


def _score_header_map(header_map: dict[int, str]) -> int:
    """Rank possible header rows by their ledger import signal."""

    detected = set(header_map.values())
    score = 0
    for column_name in (
        "posting_date",
        "account_code",
        "account_name",
        "debit_amount",
        "credit_amount",
        "debit_balance",
        "credit_balance",
        "balance",
        "balance_side",
    ):
        if column_name in detected:
            score += 3
    return score


def _canonicalize_matrix_row(
    *,
    raw_row: Sequence[str],
    header_map: dict[int, str],
) -> dict[str, str]:
    """Project one source row into canonical field names with trimmed string values."""

    canonical_row: dict[str, str] = {}
    for index, raw_value in enumerate(raw_row):
        canonical_name = header_map.get(index)
        if canonical_name is None:
            continue
        canonical_row[canonical_name] = raw_value.strip()
    return canonical_row


def _parse_gl_row(
    *,
    row: dict[str, str],
    row_number: int,
    account_code_lookup: Mapping[str, str],
) -> ImportedGeneralLedgerLineSeed:
    """Validate one canonical general-ledger row."""

    posting_date = _parse_required_date(
        row.get("posting_date"),
        field_name="posting_date",
        row_number=row_number,
    )
    account_name = _optional_text(row.get("account_name"))
    account_code = _resolve_account_code(
        row.get("account_code"),
        account_name=account_name,
        account_code_lookup=account_code_lookup,
        noun="general ledger",
        row_number=row_number,
    )
    explicit_transaction_group_value = _optional_text(row.get("transaction_group_key"))
    debit_amount, credit_amount = _resolve_import_amounts(row=row, row_number=row_number)
    return ImportedGeneralLedgerLineSeed(
        line_no=row_number - 1,
        posting_date=posting_date,
        account_code=account_code,
        account_name=account_name,
        reference=_optional_text(row.get("reference")) or explicit_transaction_group_value,
        description=_optional_text(row.get("description")),
        debit_amount=debit_amount,
        credit_amount=credit_amount,
        dimensions=_build_dimensions(row=row),
        external_ref=_optional_text(row.get("external_ref")),
        transaction_group_key=_build_transaction_group_key(
            row=row,
            posting_date=posting_date,
            line_no=row_number - 1,
        ),
    )


def _parse_tb_row(
    *,
    row: dict[str, str],
    row_number: int,
    account_code_lookup: Mapping[str, str],
) -> ImportedTrialBalanceLineSeed:
    """Validate one canonical trial-balance row."""

    account_name = _optional_text(row.get("account_name"))
    account_code = _resolve_account_code(
        row.get("account_code"),
        account_name=account_name,
        account_code_lookup=account_code_lookup,
        noun="trial balance",
        row_number=row_number,
    )
    debit_balance, credit_balance = _resolve_balance_amounts(row=row, row_number=row_number)
    return ImportedTrialBalanceLineSeed(
        line_no=row_number - 1,
        account_code=account_code,
        account_name=account_name,
        account_type=_optional_text(row.get("account_type")),
        debit_balance=debit_balance,
        credit_balance=credit_balance,
        is_active=_parse_optional_bool(row.get("is_active"), default=True),
    )


def _resolve_import_amounts(*, row: dict[str, str], row_number: int) -> tuple[Decimal, Decimal]:
    """Resolve debit/credit values from one ledger row using the supported amount schemes."""

    signed_amount = _optional_decimal(row.get("signed_amount"))
    if signed_amount is not None:
        if signed_amount == Decimal("0"):
            raise LedgerImportError(
                code=LedgerImportErrorCode.INVALID_FILE,
                message=f"Row {row_number} has a zero signed_amount; ledger rows must be non-zero.",
            )
        return (
            signed_amount if signed_amount > 0 else Decimal("0.00"),
            abs(signed_amount) if signed_amount < 0 else Decimal("0.00"),
        )

    debit_amount = _optional_decimal(row.get("debit_amount"))
    credit_amount = _optional_decimal(row.get("credit_amount"))
    if debit_amount is not None or credit_amount is not None:
        resolved_debit = debit_amount or Decimal("0.00")
        resolved_credit = credit_amount or Decimal("0.00")
        _validate_single_sided_amount(
            debit_amount=resolved_debit,
            credit_amount=resolved_credit,
            row_number=row_number,
            noun="ledger",
        )
        return resolved_debit, resolved_credit

    amount = _optional_decimal(row.get("amount"))
    line_type = _optional_text(row.get("line_type"))
    if amount is not None and line_type is not None:
        normalized_line_type = line_type.lower()
        if normalized_line_type not in {"debit", "credit"}:
            raise LedgerImportError(
                code=LedgerImportErrorCode.INVALID_FILE,
                message=(
                    f"Row {row_number} has invalid line_type {line_type!r}; "
                    "use debit or credit."
                ),
            )
        if amount <= 0:
            raise LedgerImportError(
                code=LedgerImportErrorCode.INVALID_FILE,
                message=f"Row {row_number} amount must be greater than zero.",
            )
        return (
            amount if normalized_line_type == "debit" else Decimal("0.00"),
            amount if normalized_line_type == "credit" else Decimal("0.00"),
        )

    raise LedgerImportError(
        code=LedgerImportErrorCode.INVALID_FILE,
        message=(
            f"Row {row_number} must provide either signed_amount, debit/credit amounts, "
            "or amount with line_type."
        ),
    )


def _resolve_balance_amounts(*, row: dict[str, str], row_number: int) -> tuple[Decimal, Decimal]:
    """Resolve debit/credit balances from one trial-balance row."""

    debit_balance = _optional_decimal(row.get("debit_balance"))
    credit_balance = _optional_decimal(row.get("credit_balance"))
    if debit_balance is not None or credit_balance is not None:
        resolved_debit = debit_balance or Decimal("0.00")
        resolved_credit = credit_balance or Decimal("0.00")
        _validate_single_sided_amount(
            debit_amount=resolved_debit,
            credit_amount=resolved_credit,
            row_number=row_number,
            noun="trial balance",
            allow_zero=True,
        )
        return resolved_debit, resolved_credit

    balance = _optional_decimal(row.get("balance"))
    balance_side = _optional_text(row.get("balance_side"))
    if balance is not None and balance_side is not None:
        normalized_side = balance_side.lower()
        if normalized_side not in {"debit", "credit"}:
            raise LedgerImportError(
                code=LedgerImportErrorCode.INVALID_FILE,
                message=(
                    f"Row {row_number} has invalid balance_side {balance_side!r}; "
                    "use debit or credit."
                ),
            )
        if balance < 0:
            raise LedgerImportError(
                code=LedgerImportErrorCode.INVALID_FILE,
                message=f"Row {row_number} balance must be zero or greater.",
            )
        return (
            balance if normalized_side == "debit" else Decimal("0.00"),
            balance if normalized_side == "credit" else Decimal("0.00"),
        )

    raise LedgerImportError(
        code=LedgerImportErrorCode.INVALID_FILE,
        message=(
            f"Row {row_number} must provide debit/credit balances or balance with balance_side."
        ),
    )


def _validate_single_sided_amount(
    *,
    debit_amount: Decimal,
    credit_amount: Decimal,
    row_number: int,
    noun: str,
    allow_zero: bool = False,
) -> None:
    """Ensure one row does not carry both debit and credit amounts."""

    if debit_amount < 0 or credit_amount < 0:
        raise LedgerImportError(
            code=LedgerImportErrorCode.INVALID_FILE,
            message=(
                f"Row {row_number} in the {noun} file cannot contain "
                "negative debit/credit values."
            ),
        )
    if debit_amount > 0 and credit_amount > 0:
        raise LedgerImportError(
            code=LedgerImportErrorCode.INVALID_FILE,
            message=(
                f"Row {row_number} in the {noun} file cannot contain "
                "both debit and credit values."
            ),
        )
    if not allow_zero and debit_amount == 0 and credit_amount == 0:
        raise LedgerImportError(
            code=LedgerImportErrorCode.INVALID_FILE,
            message=f"Row {row_number} in the {noun} file must contain a non-zero amount.",
        )


def _require_text(value: str | None, *, field_name: str, row_number: int) -> str:
    """Return one required non-empty text field or raise a row-specific error."""

    normalized = _optional_text(value)
    if normalized is None:
        raise LedgerImportError(
            code=LedgerImportErrorCode.INVALID_FILE,
            message=f"Row {row_number} is missing required field {field_name}.",
        )
    return normalized


def _resolve_account_code(
    value: str | None,
    *,
    account_name: str | None,
    account_code_lookup: Mapping[str, str],
    noun: str,
    row_number: int,
) -> str:
    """Resolve account identity from an explicit code or active-COA account name."""

    account_code = _optional_text(value)
    if account_code is not None:
        return account_code

    if account_name is not None:
        resolved_code = account_code_lookup.get(_normalize_account_lookup_key(account_name))
        if resolved_code is not None:
            return resolved_code

    raise LedgerImportError(
        code=LedgerImportErrorCode.INVALID_FILE,
        message=(
            f"Row {row_number} in the {noun} file is missing account_code. "
            "Upload a file with account codes, or upload/sync a COA first so account names "
            "can be resolved safely."
        ),
    )


def _optional_text(value: str | None) -> str | None:
    """Normalize optional text values and collapse blanks to null."""

    if value is None:
        return None
    normalized = value.strip()
    return normalized or None


def _optional_decimal(value: str | None) -> Decimal | None:
    """Parse one optional decimal string, preserving null when blank."""

    normalized = _optional_text(value)
    if normalized is None:
        return None
    is_parenthesized_negative = normalized.startswith("(") and normalized.endswith(")")
    sanitized = (
        normalized.strip("()")
        .translate(
            {0x2013: "-", 0x2014: "-", 0x20A6: "", 0x24: "", 0xA3: "", 0x20AC: ""}
        )
        .replace(",", "")
        .strip()
    )
    if sanitized == "-":
        return None
    try:
        amount = Decimal(sanitized)
        return -amount if is_parenthesized_negative else amount
    except InvalidOperation as error:
        raise LedgerImportError(
            code=LedgerImportErrorCode.INVALID_FILE,
            message=f"Value {normalized!r} is not a valid decimal amount.",
        ) from error


def _parse_required_date(value: str | None, *, field_name: str, row_number: int) -> date:
    """Parse one required date field from a supported spreadsheet/string format."""

    normalized = _require_text(value, field_name=field_name, row_number=row_number)
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(normalized, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(normalized).date()
    except ValueError as error:
        raise LedgerImportError(
            code=LedgerImportErrorCode.INVALID_FILE,
            message=(
                f"Row {row_number} field {field_name} must be a valid date; "
                f"received {normalized!r}."
            ),
        ) from error


def _parse_optional_bool(value: str | None, *, default: bool) -> bool:
    """Parse one optional boolean literal, returning the provided default when blank."""

    normalized = _optional_text(value)
    if normalized is None:
        return default
    lowered = normalized.lower()
    if lowered in _TRUE_LITERALS:
        return True
    if lowered in _FALSE_LITERALS:
        return False
    raise LedgerImportError(
        code=LedgerImportErrorCode.INVALID_FILE,
        message=f"Boolean field value {normalized!r} is not supported.",
    )


def _build_dimensions(*, row: dict[str, str]) -> JsonObject:
    """Extract the supported accounting-dimension fields from one import row."""

    dimensions: JsonObject = {}
    for key in ("cost_centre", "department", "project"):
        value = _optional_text(row.get(key))
        if value is not None:
            dimensions[key] = value
    return dimensions


def _is_trial_balance_summary_row(*, row: dict[str, str]) -> bool:
    """Return whether one TB row is a grand total/subtotal row rather than an account."""

    label = _optional_text(row.get("account_code")) or _optional_text(row.get("account_name"))
    if label is None:
        return True
    normalized = _normalize_account_lookup_key(label)
    return normalized in {"total", "totals", "grand_total", "grand_totals"}


def _normalize_account_code_lookup(
    account_code_lookup: Mapping[str, str] | None,
) -> dict[str, str]:
    """Normalize active-COA account name lookup keys for import resolution."""

    if account_code_lookup is None:
        return {}
    return {
        _normalize_account_lookup_key(name): str(code).strip()
        for name, code in account_code_lookup.items()
        if str(name).strip() and str(code).strip()
    }


def _normalize_account_lookup_key(value: str) -> str:
    """Normalize account names for exact-but-format-tolerant lookup."""

    return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")


def _build_transaction_group_key(
    *,
    row: dict[str, str],
    posting_date: date,
    line_no: int,
) -> str:
    """Return one canonical transaction-group key for imported GL rows."""

    explicit_group_value = _normalize_group_key_component(row.get("transaction_group_key"))
    if explicit_group_value is not None:
        return _hash_transaction_group_seed(
            posting_date=posting_date,
            source_name="explicit",
            source_value=explicit_group_value,
        )

    for source_name in ("external_ref", "reference", "description"):
        normalized_value = _normalize_group_key_component(row.get(source_name))
        if normalized_value is not None:
            return _hash_transaction_group_seed(
                posting_date=posting_date,
                source_name=source_name,
                source_value=normalized_value,
            )

    return _hash_transaction_group_seed(
        posting_date=posting_date,
        source_name="line",
        source_value=str(line_no),
    )


def _normalize_group_key_component(value: str | None) -> str | None:
    """Normalize one grouping value into a stable case-insensitive token."""

    normalized = _optional_text(value)
    if normalized is None:
        return None
    return normalized.lower()


def _hash_transaction_group_seed(
    *,
    posting_date: date,
    source_name: str,
    source_value: str,
) -> str:
    """Hash one canonical transaction grouping seed into a compact stable key."""

    seed = f"{posting_date.isoformat()}|{source_name}|{source_value}"
    return f"glgrp_{hashlib.md5(seed.encode('utf-8'), usedforsecurity=False).hexdigest()}"


def _normalize_header_name(value: str) -> str:
    """Normalize header text into a lowercase underscore form used by alias maps."""

    return (
        value.strip()
        .lower()
        .replace("-", "_")
        .replace("/", "_")
        .replace(" ", "_")
    )


def _build_import_warnings(
    *,
    document_kind: str,
    source_format: str,
    account_identity_strategy: str,
    transaction_grouping_strategy: str | None = None,
    skipped_summary_row_count: int = 0,
) -> tuple[ImportDiagnosticIssue, ...]:
    """Return non-blocking diagnostics for successfully parsed ledger imports."""

    warnings: list[ImportDiagnosticIssue] = []
    document_label = document_kind.replace("_", " ")
    if account_identity_strategy == "resolved_by_name":
        warnings.append(
            ImportDiagnosticIssue(
                code="account_names_resolved_from_active_coa",
                severity="warning",
                message=(
                    f"The {document_label} did not include account codes; account names were "
                    "resolved against the active chart of accounts."
                ),
                recovery_action=(
                    "Confirm the active COA matches the source accounting package before using "
                    "this import as a production baseline."
                ),
            )
        )
    if transaction_grouping_strategy == "derived_from_ledger_fields":
        warnings.append(
            ImportDiagnosticIssue(
                code="transaction_groups_derived",
                severity="info",
                message=(
                    "The general ledger did not include an explicit journal/transaction group "
                    "column, so grouping keys were derived from reference, description, date, "
                    "and line position."
                ),
                recovery_action=(
                    "Include voucher, journal, or transaction numbers in future GL exports when "
                    "available."
                ),
            )
        )
    if skipped_summary_row_count > 0:
        warnings.append(
            ImportDiagnosticIssue(
                code="summary_rows_skipped",
                severity="info",
                message=(
                    f"{skipped_summary_row_count} trial-balance total/subtotal row(s) were "
                    "recognized and skipped."
                ),
                recovery_action=(
                    "No action needed unless the skipped rows contain posting accounts."
                ),
            )
        )
    if source_format == "pdf":
        warnings.append(
            ImportDiagnosticIssue(
                code="searchable_pdf_table_required",
                severity="info",
                message=(
                    f"The {document_label} was parsed from searchable PDF text whose table "
                    "columns extracted cleanly."
                ),
                recovery_action=(
                    "Prefer the source XLSX/CSV export when the accounting package can provide it."
                ),
            )
        )
    return tuple(warnings)


def _resolve_import_confidence(
    *,
    source_format: str,
    account_identity_strategy: str,
    transaction_grouping_strategy: str | None,
) -> float:
    """Return a bounded parser-confidence score for successful ledger imports."""

    confidence = 0.97
    if account_identity_strategy == "resolved_by_name":
        confidence -= 0.10
    if transaction_grouping_strategy == "derived_from_ledger_fields":
        confidence -= 0.04
    if source_format == "pdf":
        confidence -= 0.06
    return max(confidence, 0.70)


def _count_account_name_resolution_rows(*, rows: Sequence[_CanonicalRow]) -> int:
    """Return how many canonical rows depend on account-name lookup."""

    return sum(
        1
        for row in rows
        if _optional_text(row.values.get("account_code")) is None
        and _optional_text(row.values.get("account_name")) is not None
    )


def _detect_gl_amount_strategy(detected_columns: frozenset[str]) -> str:
    """Return the amount scheme detected from GL headers."""

    if "signed_amount" in detected_columns:
        return "signed_amount"
    if {"debit_amount", "credit_amount"}.intersection(detected_columns):
        return "debit_credit_columns"
    if "amount" in detected_columns and "line_type" in detected_columns:
        return "amount_with_line_type"
    return "row_level_amount_resolution"


def _detect_tb_balance_strategy(detected_columns: frozenset[str]) -> str:
    """Return the balance scheme detected from trial-balance headers."""

    if {"debit_balance", "credit_balance"}.intersection(detected_columns):
        return "debit_credit_balance_columns"
    if "balance" in detected_columns and "balance_side" in detected_columns:
        return "balance_with_side"
    return "row_level_balance_resolution"


__all__ = [
    "ImportedGeneralLedgerFile",
    "ImportedGeneralLedgerLineSeed",
    "ImportedTrialBalanceFile",
    "ImportedTrialBalanceLineSeed",
    "LedgerImportError",
    "LedgerImportErrorCode",
    "import_general_ledger_file",
    "import_trial_balance_file",
]
