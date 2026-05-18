"""
Purpose: Parse and validate chart-of-accounts upload files.
Scope: CSV/XLSX/searchable-PDF decoding, header normalization, account row validation,
duplicate detection, parent-link validation, and import metadata generation.
Dependencies: Python CSV/io helpers, openpyxl, and shared JSON type aliases.
"""

from __future__ import annotations

import csv
from collections.abc import Sequence
from dataclasses import dataclass
from enum import StrEnum
from io import BytesIO, StringIO
from pathlib import Path
from typing import Final

from openpyxl import load_workbook  # type: ignore[import-untyped]
from pypdf import PdfReader
from pypdf.errors import PdfReadError
from services.common.types import JsonObject


class CoaImportErrorCode(StrEnum):
    """Enumerate stable validation codes surfaced by COA file imports."""

    INVALID_FILE = "invalid_file"
    UNSUPPORTED_FILE_TYPE = "unsupported_file_type"


class CoaImportError(ValueError):
    """Represent a fail-fast COA import validation failure."""

    def __init__(self, *, code: CoaImportErrorCode, message: str) -> None:
        """Capture a stable validation code and operator-facing message."""

        super().__init__(message)
        self.code = code
        self.message = message


@dataclass(frozen=True, slots=True)
class ImportedCoaAccountSeed:
    """Describe one validated account row parsed from an upload file."""

    account_code: str
    account_name: str
    account_type: str
    parent_account_code: str | None
    is_postable: bool
    is_active: bool
    external_ref: str | None
    dimension_defaults: JsonObject


@dataclass(frozen=True, slots=True)
class ImportedCoaFile:
    """Describe the fully validated COA upload payload returned to service workflows."""

    accounts: tuple[ImportedCoaAccountSeed, ...]
    import_metadata: JsonObject


_REQUIRED_COLUMNS: Final[frozenset[str]] = frozenset(
    {"account_code", "account_name", "account_type"}
)

_HEADER_ALIASES: Final[dict[str, str]] = {
    "account": "account_name",
    "account_code": "account_code",
    "account_name": "account_name",
    "account_title": "account_name",
    "account_number": "account_code",
    "account_type": "account_type",
    "active": "is_active",
    "category": "account_type",
    "code": "account_code",
    "cost_centre": "cost_centre",
    "cost_center": "cost_centre",
    "default_cost_centre": "cost_centre",
    "default_cost_center": "cost_centre",
    "default_department": "department",
    "default_project": "project",
    "department": "department",
    "external_ref": "external_ref",
    "external_reference": "external_ref",
    "financial_statement": "financial_statement",
    "fs": "financial_statement",
    "gl_code": "account_code",
    "group": "account_group",
    "is_active": "is_active",
    "is_postable": "is_postable",
    "name": "account_name",
    "normal_balance": "normal_balance",
    "normally": "normal_balance",
    "parent": "parent_account_code",
    "parent_account": "parent_account_code",
    "parent_account_code": "parent_account_code",
    "parent_code": "parent_account_code",
    "postable": "is_postable",
    "project": "project",
    "qbo_id": "external_ref",
    "quickbooks_id": "external_ref",
    "statement": "financial_statement",
    "sub_group": "account_sub_group",
    "subgroup": "account_sub_group",
    "type": "account_type",
}

_TRUE_LITERALS: Final[frozenset[str]] = frozenset({"1", "t", "true", "y", "yes"})
_FALSE_LITERALS: Final[frozenset[str]] = frozenset({"0", "f", "false", "n", "no"})


def import_coa_file(*, filename: str, payload: bytes) -> ImportedCoaFile:
    """Parse a CSV/XLSX/searchable-PDF COA payload into canonical account seeds."""

    if not payload:
        raise CoaImportError(
            code=CoaImportErrorCode.INVALID_FILE,
            message="Uploaded COA files cannot be empty.",
        )

    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        rows, detected_columns = _read_csv_rows(payload=payload)
        source_format = "csv"
    elif suffix in {".xlsx", ".xlsm"}:
        rows, detected_columns = _read_workbook_rows(payload=payload)
        source_format = "xlsx"
    elif suffix == ".pdf":
        rows, detected_columns = _read_pdf_rows(payload=payload)
        source_format = "pdf"
    else:
        raise CoaImportError(
            code=CoaImportErrorCode.UNSUPPORTED_FILE_TYPE,
            message="Upload a CSV, XLSX, or searchable PDF chart-of-accounts file.",
        )

    if not rows:
        raise CoaImportError(
            code=CoaImportErrorCode.INVALID_FILE,
            message="The COA file does not contain any account rows.",
        )

    accounts = tuple(
        _parse_account_row(row=row, row_number=index + 2) for index, row in enumerate(rows)
    )
    _validate_accounts(accounts)

    metadata: JsonObject = {
        "detected_columns": ", ".join(sorted(detected_columns)),
        "format": source_format,
        "row_count": len(accounts),
        "uploaded_filename": filename,
    }
    return ImportedCoaFile(accounts=accounts, import_metadata=metadata)


def _read_csv_rows(*, payload: bytes) -> tuple[tuple[dict[str, str], ...], frozenset[str]]:
    """Read CSV rows and normalize header names into canonical COA column keys."""

    try:
        decoded_payload = payload.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise CoaImportError(
            code=CoaImportErrorCode.INVALID_FILE,
            message="CSV files must be UTF-8 encoded.",
        ) from error

    reader = csv.reader(StringIO(decoded_payload))
    raw_rows = tuple(tuple(cell.strip() for cell in row) for row in reader)
    if not raw_rows:
        raise CoaImportError(
            code=CoaImportErrorCode.INVALID_FILE,
            message="CSV files must include a header row.",
        )

    return _read_matrix_rows(raw_rows=raw_rows)


def _read_workbook_rows(*, payload: bytes) -> tuple[tuple[dict[str, str], ...], frozenset[str]]:
    """Read the first worksheet in an XLSX payload using normalized COA headers."""

    try:
        workbook = load_workbook(filename=BytesIO(payload), read_only=True, data_only=True)
    except Exception as error:
        raise CoaImportError(
            code=CoaImportErrorCode.INVALID_FILE,
            message="The workbook could not be opened. Upload a valid XLSX file.",
        ) from error

    worksheet = workbook.active
    raw_rows = tuple(
        tuple("" if cell is None else str(cell).strip() for cell in row)
        for row in worksheet.iter_rows(values_only=True)
    )
    if not raw_rows:
        raise CoaImportError(
            code=CoaImportErrorCode.INVALID_FILE,
            message="The workbook does not contain a header row.",
        )

    return _read_matrix_rows(raw_rows=raw_rows)


def _read_pdf_rows(*, payload: bytes) -> tuple[tuple[dict[str, str], ...], frozenset[str]]:
    """Read searchable PDF text when the table structure is preserved by extraction."""

    try:
        reader = PdfReader(BytesIO(payload))
    except PdfReadError as error:
        raise CoaImportError(
            code=CoaImportErrorCode.INVALID_FILE,
            message="The PDF could not be opened. Upload a valid searchable PDF.",
        ) from error

    raw_rows = tuple(
        tuple(cell.strip() for cell in line.split("|"))
        for page in reader.pages
        for line in (page.extract_text() or "").splitlines()
        if "|" in line
    )
    if not raw_rows:
        raise CoaImportError(
            code=CoaImportErrorCode.INVALID_FILE,
            message=(
                "The PDF text did not preserve a chart-of-accounts table. Upload the source "
                "spreadsheet as XLSX/CSV, or a searchable PDF whose table columns extract cleanly."
            ),
        )

    return _read_matrix_rows(raw_rows=raw_rows)


def _read_matrix_rows(
    *,
    raw_rows: Sequence[Sequence[str]],
) -> tuple[tuple[dict[str, str], ...], frozenset[str]]:
    """Find the header row in a worksheet/CSV matrix and normalize following rows."""

    non_empty_rows = tuple(row for row in raw_rows if any(cell.strip() for cell in row))
    if not non_empty_rows:
        raise CoaImportError(
            code=CoaImportErrorCode.INVALID_FILE,
            message="The COA file does not contain any account rows.",
        )

    best_header: tuple[int, dict[int, str]] | None = None
    best_score = -1
    for index, row in enumerate(non_empty_rows[:25]):
        header_map = _build_header_map(row, validate=False)
        score = _score_header_map(header_map)
        if score > best_score:
            best_header = (index, header_map)
            best_score = score

    if best_header is None:
        _raise_missing_header_error(())
    assert best_header is not None

    header_index, canonical_header_map = best_header
    _validate_header_map(tuple(canonical_header_map.values()))
    headers = tuple(non_empty_rows[header_index])
    rows: list[dict[str, str]] = []
    for raw_row in non_empty_rows[header_index + 1 :]:
        canonical = _canonicalize_matrix_row(
            raw_row=raw_row,
            headers=headers,
            header_map=canonical_header_map,
        )
        if canonical:
            rows.append(canonical)

    return tuple(rows), frozenset(canonical_header_map.values())


def _build_header_map(headers: Sequence[str], *, validate: bool = True) -> dict[int, str]:
    """Map source header names to canonical COA field names and validate required columns."""

    header_map: dict[int, str] = {}
    for index, header in enumerate(headers):
        normalized = _normalize_header_name(header)
        canonical_name = _HEADER_ALIASES.get(normalized)
        if canonical_name is None:
            continue
        header_map[index] = canonical_name

    if validate:
        _validate_header_map(tuple(header_map.values()))

    return header_map


def _validate_header_map(detected_columns: Sequence[str]) -> None:
    """Ensure headers include enough account identity and type evidence."""

    detected = set(detected_columns)
    missing = sorted({"account_code", "account_name"}.difference(detected))
    has_type_evidence = bool(
        {"account_type", "financial_statement", "account_group", "account_sub_group"}.intersection(
            detected
        )
    )
    if "account_type" not in detected and not has_type_evidence:
        missing.append("account_type")
    if missing:
        _raise_missing_header_error(missing)


def _raise_missing_header_error(missing: Sequence[str]) -> None:
    """Raise the canonical missing-header import error."""

    missing_columns = ", ".join(missing) if missing else ", ".join(sorted(_REQUIRED_COLUMNS))
    raise CoaImportError(
        code=CoaImportErrorCode.INVALID_FILE,
        message=f"The COA file is missing required columns: {missing_columns}.",
    )


def _score_header_map(header_map: dict[int, str]) -> int:
    """Rank a candidate header row by how much COA evidence it contains."""

    detected = set(header_map.values())
    score = 0
    for column_name in ("account_code", "account_name", "account_type"):
        if column_name in detected:
            score += 4
    for column_name in ("financial_statement", "account_group", "account_sub_group"):
        if column_name in detected:
            score += 2
    return score


def _canonicalize_matrix_row(
    *,
    raw_row: Sequence[str],
    headers: Sequence[str],
    header_map: dict[int, str],
) -> dict[str, str]:
    """Project one source row into canonical field names with raw string values."""

    canonical_row: dict[str, str] = {}
    del headers
    for index, value in enumerate(raw_row):
        canonical_name = header_map.get(index)
        if canonical_name is None:
            continue
        canonical_row[canonical_name] = value

    return canonical_row


def _parse_account_row(*, row: dict[str, str], row_number: int) -> ImportedCoaAccountSeed:
    """Validate one canonical row dictionary and convert it into an account seed."""

    account_code = _require_text(
        row.get("account_code"), field_name="account_code", row_number=row_number
    )
    account_name = _require_text(
        row.get("account_name"), field_name="account_name", row_number=row_number
    )
    raw_account_type = _normalize_optional_text(row.get("account_type"))
    account_type = (
        _normalize_account_type(raw_account_type)
        if raw_account_type is not None
        else _infer_account_type(row=row, row_number=row_number)
    )
    parent_account_code = _normalize_optional_text(row.get("parent_account_code"))
    external_ref = _normalize_optional_text(row.get("external_ref"))

    return ImportedCoaAccountSeed(
        account_code=account_code,
        account_name=account_name,
        account_type=account_type,
        parent_account_code=parent_account_code,
        is_postable=_parse_boolean(
            row.get("is_postable"), default=True, field_name="is_postable", row_number=row_number
        ),
        is_active=_parse_boolean(
            row.get("is_active"), default=True, field_name="is_active", row_number=row_number
        ),
        external_ref=external_ref,
        dimension_defaults=_build_dimension_defaults(row),
    )


def _validate_accounts(accounts: tuple[ImportedCoaAccountSeed, ...]) -> None:
    """Run duplicate-code and parent-link checks across the parsed account list."""

    codes = [account.account_code for account in accounts]
    unique_codes = set(codes)
    if len(codes) != len(unique_codes):
        duplicated = sorted({code for code in codes if codes.count(code) > 1})
        duplicate_codes = ", ".join(duplicated)
        raise CoaImportError(
            code=CoaImportErrorCode.INVALID_FILE,
            message=f"Duplicate account_code values were found: {duplicate_codes}.",
        )

    for account in accounts:
        parent_code = account.parent_account_code
        if parent_code is None:
            continue
        if parent_code == account.account_code:
            raise CoaImportError(
                code=CoaImportErrorCode.INVALID_FILE,
                message=(
                    f"Account {account.account_code} cannot reference itself as "
                    "parent_account_code."
                ),
            )
        if parent_code not in unique_codes:
            raise CoaImportError(
                code=CoaImportErrorCode.INVALID_FILE,
                message=(
                    f"Account {account.account_code} references unknown parent_account_code "
                    f"{parent_code}."
                ),
            )


def _build_dimension_defaults(row: dict[str, str]) -> JsonObject:
    """Build optional dimension defaults from known COA upload columns."""

    defaults: JsonObject = {}
    for source_field, target_key in (
        ("cost_centre", "cost_centre"),
        ("department", "department"),
        ("project", "project"),
    ):
        value = _normalize_optional_text(row.get(source_field))
        if value is not None:
            defaults[target_key] = value

    return defaults


def _normalize_header_name(value: str) -> str:
    """Normalize a file header to a lower snake_case comparison key."""

    return value.strip().lower().replace("-", "_").replace(" ", "_")


def _require_text(value: str | None, *, field_name: str, row_number: int) -> str:
    """Return required text field values or raise an import error with row context."""

    normalized = _normalize_optional_text(value)
    if normalized is not None:
        return normalized

    raise CoaImportError(
        code=CoaImportErrorCode.INVALID_FILE,
        message=f"Row {row_number} is missing required field: {field_name}.",
    )


def _normalize_optional_text(value: str | None) -> str | None:
    """Trim optional text values and collapse blanks to null."""

    if value is None:
        return None

    normalized = value.strip()
    return normalized or None


def _normalize_account_type(value: str) -> str:
    """Normalize account-type labels to lower snake_case values."""

    normalized = _normalize_accounting_label(value)
    aliases = {
        "assets": "asset",
        "asset": "asset",
        "current_assets": "asset",
        "fixed_assets": "asset",
        "long_term_assets": "asset",
        "non_current_assets": "asset",
        "liabilities": "liability",
        "liability": "liability",
        "current_liabilities": "liability",
        "long_term_liabilities": "liability",
        "non_current_liabilities": "liability",
        "equity": "equity",
        "owners_equity": "equity",
        "owner_s_equity": "equity",
        "capital": "equity",
        "income": "revenue",
        "revenue": "revenue",
        "sales": "revenue",
        "other_income": "other_income",
        "cost_of_sales": "cost_of_sales",
        "cost_of_goods_sold": "cost_of_sales",
        "cogs": "cost_of_sales",
        "expenses": "expense",
        "expense": "expense",
        "operating_expenses": "expense",
        "other_expenses": "other_expense",
        "other_expense": "other_expense",
    }
    return aliases.get(normalized, normalized)


def _infer_account_type(*, row: dict[str, str], row_number: int) -> str:
    """Infer account type from real-world COA grouping columns."""

    signals = " ".join(
        value
        for key in ("account_group", "account_sub_group", "financial_statement", "account_name")
        if (value := _normalize_optional_text(row.get(key))) is not None
    )
    normalized = _normalize_accounting_label(signals)
    if not normalized:
        raise CoaImportError(
            code=CoaImportErrorCode.INVALID_FILE,
            message=(
                f"Row {row_number} is missing account_type and does not include enough "
                "financial statement/group evidence to infer it."
            ),
        )

    if any(token in normalized for token in ("asset", "cash", "receivable", "inventory")):
        return "asset"
    if any(token in normalized for token in ("liabil", "payable", "creditor", "loan")):
        return "liability"
    if any(token in normalized for token in ("equity", "capital", "retained_earning")):
        return "equity"
    if any(token in normalized for token in ("cost_of_sales", "cost_of_goods", "cogs")):
        return "cost_of_sales"
    if any(token in normalized for token in ("expense", "wage", "salary", "rent", "utility")):
        return "expense"
    if any(token in normalized for token in ("revenue", "income", "sales", "service")):
        return "revenue"

    raise CoaImportError(
        code=CoaImportErrorCode.INVALID_FILE,
        message=(
            f"Row {row_number} is missing account_type and the importer could not infer one "
            "from the financial statement/group columns."
        ),
    )


def _normalize_accounting_label(value: str) -> str:
    """Normalize account labels into the comparison surface used by format inference."""

    return (
        value.strip()
        .lower()
        .replace("&", " and ")
        .replace("/", " ")
        .replace("-", " ")
        .replace("'", " ")
        .replace(" ", "_")
    )


def _parse_boolean(
    value: str | None,
    *,
    default: bool,
    field_name: str,
    row_number: int,
) -> bool:
    """Parse optional boolean literals with explicit row-scoped validation errors."""

    normalized = _normalize_optional_text(value)
    if normalized is None:
        return default

    lowered = normalized.lower()
    if lowered in _TRUE_LITERALS:
        return True
    if lowered in _FALSE_LITERALS:
        return False

    raise CoaImportError(
        code=CoaImportErrorCode.INVALID_FILE,
        message=(f"Row {row_number} has invalid boolean value for {field_name}: {normalized}."),
    )


__all__ = [
    "CoaImportError",
    "CoaImportErrorCode",
    "ImportedCoaAccountSeed",
    "ImportedCoaFile",
    "import_coa_file",
]
