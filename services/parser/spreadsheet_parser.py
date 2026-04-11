"""
Purpose: Parse Excel workbooks and CSV files into normalized deterministic table payloads.
Scope: CSV dialect detection, XLSX worksheet extraction, row/header normalization,
text flattening, split/group metadata, and JSON normalized derivatives.
Dependencies: Python CSV/JSON libraries, openpyxl, and parser-domain models.
"""

from __future__ import annotations

import csv
import json
from collections.abc import Iterable, Sequence
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO, StringIO
from pathlib import PurePath

from openpyxl import load_workbook  # type: ignore[import-untyped]
from services.parser.document_splitter import detect_document_splits
from services.parser.models import (
    ParsedPage,
    ParsedTable,
    ParserErrorCode,
    ParserPipelineError,
    ParserResult,
)

CSV_MIME_TYPES = {"text/csv", "application/csv"}
EXCEL_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}


def parse_spreadsheet_document(*, payload: bytes, filename: str, mime_type: str) -> ParserResult:
    """Parse a CSV or Excel source document into normalized table and text outputs."""

    normalized_mime_type = mime_type.strip().lower()
    if normalized_mime_type in CSV_MIME_TYPES or filename.lower().endswith(".csv"):
        return _parse_csv(payload=payload, filename=filename)

    if normalized_mime_type in EXCEL_MIME_TYPES or filename.lower().endswith((".xlsx", ".xlsm")):
        return _parse_excel(payload=payload, filename=filename)

    raise ParserPipelineError(
        code=ParserErrorCode.UNSUPPORTED_MIME,
        message=f"Spreadsheet parser does not support MIME type {mime_type}.",
    )


def _parse_csv(*, payload: bytes, filename: str) -> ParserResult:
    """Parse CSV bytes using deterministic dialect sniffing and header normalization."""

    text = _decode_text_payload(payload=payload, filename=filename)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(StringIO(text), dialect=dialect)
    raw_rows = tuple(tuple(cell.strip() for cell in row) for row in reader)
    table = _rows_to_table(
        table_name=PurePath(filename).stem or "csv",
        raw_rows=raw_rows,
        source_ref={"kind": "csv", "filename": filename},
    )
    parse_text = _render_tables_as_text((table,))
    result = ParserResult(
        parser_name="spreadsheet.csv",
        text=parse_text,
        pages=(
            ParsedPage(page_number=1, text=parse_text, extraction_method="csv"),
        ),
        tables=(table,),
        page_count=1,
        metadata={
            "source_format": "csv",
            "filename": filename,
            "row_count": len(table.rows),
            "table_count": 1,
        },
        normalized_filename=f"{PurePath(filename).stem or 'csv'}-normalized.json",
        normalized_content_type="application/json",
    )
    result.split_candidates = detect_document_splits(
        pages=result.pages,
        tables=result.tables,
        filename=filename,
    )
    result.set_normalized_payload(_serialize_normalized_payload(result))
    return result


def _parse_excel(*, payload: bytes, filename: str) -> ParserResult:
    """Parse an Excel workbook into one normalized table per non-empty worksheet."""

    try:
        workbook = load_workbook(filename=BytesIO(payload), data_only=True, read_only=True)
    except Exception as error:
        raise ParserPipelineError(
            code=ParserErrorCode.PARSE_FAILED,
            message=f"{filename} could not be read as an Excel workbook.",
        ) from error

    tables: list[ParsedTable] = []
    pages: list[ParsedPage] = []
    for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
        raw_rows = tuple(
            tuple(_format_cell_value(cell) for cell in row)
            for row in worksheet.iter_rows(values_only=True)
        )
        table = _rows_to_table(
            table_name=worksheet.title,
            raw_rows=raw_rows,
            source_ref={"kind": "worksheet", "sheet_name": worksheet.title},
        )
        if not table.columns and not table.rows:
            continue

        tables.append(table)
        pages.append(
            ParsedPage(
                page_number=sheet_index,
                text=_render_tables_as_text((table,)),
                extraction_method="spreadsheet",
            )
        )

    if not tables:
        raise ParserPipelineError(
            code=ParserErrorCode.PARSE_FAILED,
            message=f"{filename} did not contain any non-empty worksheets.",
        )

    parse_text = _render_tables_as_text(tuple(tables))
    result = ParserResult(
        parser_name="spreadsheet.excel",
        text=parse_text,
        pages=tuple(pages),
        tables=tuple(tables),
        page_count=len(tables),
        metadata={
            "source_format": "excel",
            "filename": filename,
            "sheet_count": len(workbook.worksheets),
            "table_count": len(tables),
        },
        normalized_filename=f"{PurePath(filename).stem or 'workbook'}-normalized.json",
        normalized_content_type="application/json",
    )
    result.split_candidates = detect_document_splits(
        pages=result.pages,
        tables=result.tables,
        filename=filename,
    )
    result.set_normalized_payload(_serialize_normalized_payload(result))
    return result


def _rows_to_table(
    *,
    table_name: str,
    raw_rows: Sequence[Sequence[str]],
    source_ref: dict[str, str],
) -> ParsedTable:
    """Convert raw row values into a table with stable headers and row dictionaries."""

    non_empty_rows = tuple(row for row in raw_rows if any(cell.strip() for cell in row))
    if not non_empty_rows:
        return ParsedTable(name=table_name, columns=(), rows=(), source_ref=dict(source_ref))

    width = max(len(row) for row in non_empty_rows)
    header_row = _pad_row(non_empty_rows[0], width)
    columns = _normalize_headers(header_row)
    data_rows: list[dict[str, str]] = []
    for row_number, raw_row in enumerate(non_empty_rows[1:], start=2):
        padded_row = _pad_row(raw_row, width)
        row_payload = {
            columns[index]: value
            for index, value in enumerate(padded_row)
            if value or columns[index].startswith("column_")
        }
        row_payload["source_row_number"] = str(row_number)
        data_rows.append(row_payload)

    return ParsedTable(
        name=table_name,
        columns=columns,
        rows=tuple(data_rows),
        source_ref=dict(source_ref),
    )


def _decode_text_payload(*, payload: bytes, filename: str) -> str:
    """Decode CSV bytes using the canonical accepted encodings."""

    for encoding in ("utf-8-sig", "utf-8"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue

    raise ParserPipelineError(
        code=ParserErrorCode.PARSE_FAILED,
        message=f"{filename} must be UTF-8 encoded CSV for the local demo parser.",
    )


def _format_cell_value(value: object) -> str:
    """Render an Excel cell value into a JSON-safe deterministic string."""

    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)

    return str(value).strip()


def _pad_row(row: Sequence[str], width: int) -> tuple[str, ...]:
    """Pad a row to the table width so headers and values align deterministically."""

    return tuple(row[index].strip() if index < len(row) else "" for index in range(width))


def _normalize_headers(header_row: Sequence[str]) -> tuple[str, ...]:
    """Normalize duplicate or blank spreadsheet headers into stable column names."""

    columns: list[str] = []
    seen_counts: dict[str, int] = {}
    for index, raw_header in enumerate(header_row, start=1):
        base_header = raw_header.strip() or f"column_{index}"
        key = base_header
        if key in seen_counts:
            seen_counts[key] += 1
            key = f"{base_header}_{seen_counts[base_header]}"
        else:
            seen_counts[key] = 1
        columns.append(key)

    return tuple(columns)


def _render_tables_as_text(tables: Iterable[ParsedTable]) -> str:
    """Flatten parsed tables into reviewable text while keeping row boundaries visible."""

    blocks: list[str] = []
    for table in tables:
        blocks.append(f"Table: {table.name}")
        if table.columns:
            blocks.append(" | ".join(table.columns))
        for row in table.rows:
            blocks.append(" | ".join(row.get(column, "") for column in table.columns))

    return "\n".join(blocks).strip()


def _serialize_normalized_payload(result: ParserResult) -> bytes:
    """Serialize normalized spreadsheet output as deterministic UTF-8 JSON bytes."""

    return json.dumps(
        result.raw_parse_payload(),
        ensure_ascii=True,
        indent=2,
        sort_keys=True,
    ).encode("utf-8")


__all__ = ["parse_spreadsheet_document"]
